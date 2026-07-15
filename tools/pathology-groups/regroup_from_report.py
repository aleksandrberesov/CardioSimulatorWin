#!/usr/bin/env python3
"""Regroup a CardioSimulator dataset ZIP using the Combined Pathologies Report as the group source.

The big dataset (``Pathologies.All.corrected.grouped.enumerated.fixed.zip``) stores every
pathology's title as a ``" + "``-joined list of condition names, e.g.::

    Sinus Bradycardia + Left ventricular hypertrophy + ST segment changes

The report Excel (``Combined_Pathologies_Report_YYYY-MM-DD.xlsx``, sheet
``Справочник состояний в1``) is the customer's authoritative condition -> group table:
each condition has an English name and a Russian group label. This tool makes that table
the single source of grouping:

  1. Read the Excel sheet -> map condition (English name, lower-cased) -> group KEY
     (via the Russian label -> key table below), plus ``condition_aliases.tsv`` for
     title wording that is not verbatim in the Excel (synonyms, typos, raw SNOMED codes).
  2. Split every manifest title on ``" + "`` into component conditions; map each to a key.
  3. Pick ONE group per pathology by a configurable clinical-severity priority
     (multi-condition records take the highest-priority group among their components).
  4. Rewrite the ``group:`` field in ``manifest.txt`` and in every ``.dat`` header, in place.
  5. A record whose title has NO mappable component keeps its EXISTING group (hand-curated
     built-ins and one-off legacy titles are never downgraded); every such case is reported.

By default the run is a DRY RUN: it reads only ``manifest.txt`` (fast), prints the old->new
group distribution and writes a per-record change report (TSV). It never touches ``--in``.
Pass ``--out`` to additionally stream a full regrouped ZIP (rewrites the header of each
``.dat`` and copies the sample data). ``--out`` must differ from ``--in``.

Requires ``openpyxl`` (``pip install openpyxl``) only to read the .xlsx.

Examples
--------
  # analyse only (seconds): prints impact, writes regroup-changes.tsv next to the report
  python tools/pathology-groups/regroup_from_report.py \
      --in  E:/VLN_Project/Data/Pathologies.All.corrected.grouped.enumerated.fixed.zip \
      --report E:/VLN_Project/Data/Combined_Pathologies_Report_2026-07-10.xlsx

  # also emit the regrouped dataset
  python tools/pathology-groups/regroup_from_report.py \
      --in  .../fixed.zip --report .../report.xlsx \
      --out E:/VLN_Project/Data/Pathologies.All.regrouped.zip
"""
import argparse
import collections
import os
import sys
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_ALIASES = os.path.join(HERE, "condition_aliases.tsv")
DEFAULT_SHEET = "Справочник состояний в1"

# Russian group label (as written in the report) -> group key used in groups.txt / manifest.
# The report uses short labels; groups.txt uses long display names. Keys are the stable link.
LABEL_TO_KEY = {
    "синусовый": "sinus",
    "нарушения ритма": "arrhythmia",
    "нарушения проводимости": "conduction",
    "гипертрофия": "hypertrophy",
    "ишемия": "ischemia",
    "инфаркт": "infarction",
    "электролитные/токсические": "electrolyte",
    "синдромы/каналопатии": "syndromes",
    "экс": "pacemaker",
    "особые/ось/вольтаж": "special",
    # patient-context groups (not condition-based; here only so an explicit label resolves)
    "педиатрия": "pediatric",
    "новорождённые": "newborn",
    "беременные": "pregnant",
    "клинические случаи": "clinical",
}

# Tie-break for multi-condition pathologies: the highest-priority group present wins.
# This mirrors the established convention in update_zip_groups.py (which the current
# dataset already follows ~71% of the time), with pacemaker inserted after conduction.
# Override at the CLI with --priority a,b,c,...
DEFAULT_PRIORITY = [
    "infarction",
    "ischemia",
    "electrolyte",
    "syndromes",
    "conduction",
    "pacemaker",
    "arrhythmia",
    "hypertrophy",
    "sinus",
    "special",
]


def norm(s):
    """Lower-case + collapse internal whitespace for robust matching."""
    return " ".join(str(s).split()).lower()


# --------------------------------------------------------------------------- #
# Catalog / map loading
# --------------------------------------------------------------------------- #
def read_group_keys_from_zip(zf):
    """Return the ordered list of group keys declared in the zip's groups.txt (if any)."""
    keys = []
    if "groups.txt" in zf.namelist():
        text = zf.read("groups.txt").decode("utf-8-sig")
        for raw in text.split("\n"):
            line = raw.strip()
            if line.startswith("group:"):
                keys.append(line.split(";", 1)[0][len("group:"):].strip())
    return keys


def load_excel_condition_map(xlsx_path, sheet, label_to_key):
    """Read the report sheet -> {condition_name_lower: group_key}. Returns (map, warnings)."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required to read the .xlsx report. "
              "Install it with: pip install openpyxl", file=sys.stderr)
        raise SystemExit(2)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        print(f"ERROR: sheet {sheet!r} not found. Sheets: {wb.sheetnames}", file=sys.stderr)
        raise SystemExit(2)
    ws = wb[sheet]

    rows = list(ws.iter_rows(values_only=True))
    # Locate the columns by header text (robust to column reordering).
    header = [norm(c) if c is not None else "" for c in rows[0]]
    def col(*names):
        for i, h in enumerate(header):
            if h in names:
                return i
        return None
    c_en = col("название (en)", "name (en)", "en")
    c_grp = col("группа", "group")
    if c_en is None or c_grp is None:
        # Fall back to the known layout: Акроним | EN | RU | Группа | SNOMED | count
        c_en, c_grp = 1, 3

    cond_map = {}
    warnings = []
    unknown_labels = set()
    for r in rows[1:]:
        if r is None:
            continue
        en = r[c_en] if c_en < len(r) else None
        grp = r[c_grp] if c_grp < len(r) else None
        if not en or not grp:
            continue
        key = label_to_key.get(norm(grp))
        if key is None:
            unknown_labels.add(str(grp).strip())
            continue
        cond_map[norm(en)] = key
    for lbl in sorted(unknown_labels):
        warnings.append(f"report group label not mapped to a key (skipped): {lbl!r}")
    return cond_map, warnings


def load_aliases(path):
    """Read condition_aliases.tsv -> {component_lower: group_key}."""
    aliases = {}
    if not path or not os.path.isfile(path):
        return aliases
    with open(path, encoding="utf-8-sig") as f:
        for raw in f:
            line = raw.rstrip("\n")
            if not line.strip() or line.lstrip().startswith("#"):
                continue
            parts = line.split("\t")
            if len(parts) < 2:
                parts = line.split()  # tolerate space-separated
            if len(parts) < 2:
                continue
            comp = norm(parts[0])
            key = parts[-1].strip()
            aliases[comp] = key
    return aliases


# --------------------------------------------------------------------------- #
# Grouping logic
# --------------------------------------------------------------------------- #
def resolve_group(title, cond_map, aliases, priority):
    """Return (group_key or None, matched_keys, unmatched_components) for a title."""
    matched = []
    unmatched = []
    for part in title.split(" + "):
        p = norm(part)
        if not p:
            continue
        key = cond_map.get(p) or aliases.get(p)
        if key:
            matched.append(key)
        else:
            unmatched.append(part.strip())
    if not matched:
        return None, [], unmatched
    present = set(matched)
    for g in priority:
        if g in present:
            return g, matched, unmatched
    return matched[0], matched, unmatched  # keys present but none in priority list


# --------------------------------------------------------------------------- #
# Manifest / .dat rewriting (in place, preserving field order)
# --------------------------------------------------------------------------- #
def set_manifest_group(line, value):
    """Replace the group: field of a manifest pathology line in place (or insert after title)."""
    parts = line.split(";")
    for i, p in enumerate(parts):
        if p.startswith("group:"):
            parts[i] = "group:" + value
            return ";".join(parts)
    # insert after the title: field if present, else append
    for i, p in enumerate(parts):
        if p.startswith("title:"):
            parts.insert(i + 1, "group:" + value)
            return ";".join(parts)
    parts.append("group:" + value)
    return ";".join(parts)


def set_dat_group(text, value):
    """Replace the group: line in a .dat header (before the first blank line) in place."""
    lines = text.split("\n")
    for i, l in enumerate(lines):
        if l == "":
            break
        if l.startswith("group:"):
            lines[i] = "group:" + value
            return "\n".join(lines)
    for i, l in enumerate(lines):
        if l.startswith("pathology:"):
            lines.insert(i + 1, "group:" + value)
            return "\n".join(lines)
    lines.insert(0, "group:" + value)
    return "\n".join(lines)


def parse_manifest(text):
    """Yield (raw_line_index, id, current_group, title) for each pathology line."""
    for idx, line in enumerate(text.split("\n")):
        if line.startswith("pathology:") and not line.lstrip().startswith("#"):
            d = {}
            parts = line.split(";")
            pid = parts[0].split(":", 1)[1].strip()
            for p in parts[1:]:
                if ":" in p:
                    k, v = p.split(":", 1)
                    d[k.strip()] = v
            yield idx, pid, d.get("group") or None, d.get("title", "")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(
        description="Regroup a CardioSimulator dataset ZIP using the report Excel as the group source.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--in", dest="inp", required=True, help="input dataset ZIP")
    ap.add_argument("--report", required=True, help="Combined Pathologies Report .xlsx")
    ap.add_argument("--sheet", default=DEFAULT_SHEET, help=f"condition sheet (default: {DEFAULT_SHEET!r})")
    ap.add_argument("--aliases", default=DEFAULT_ALIASES, help="condition alias TSV")
    ap.add_argument("--priority", default=",".join(DEFAULT_PRIORITY),
                    help="comma-separated tie-break priority (highest first)")
    ap.add_argument("--out", default=None, help="output ZIP (omit for dry-run). Must differ from --in")
    ap.add_argument("--report-out", default=None,
                    help="per-record change report TSV (default: regroup-changes.tsv next to --report)")
    ap.add_argument("--compresslevel", type=int, default=6, help="deflate level for --out (0-9, default 6)")
    args = ap.parse_args()

    if not os.path.isfile(args.inp):
        print(f"ERROR: input zip not found: {args.inp}", file=sys.stderr)
        return 2
    if args.out and os.path.abspath(args.out) == os.path.abspath(args.inp):
        print("ERROR: --out must differ from --in (never overwrites the source)", file=sys.stderr)
        return 2

    priority = [g.strip() for g in args.priority.split(",") if g.strip()]

    # 1. Build the condition -> group map from the report + aliases.
    cond_map, warnings = load_excel_condition_map(args.report, args.sheet, LABEL_TO_KEY)
    aliases = load_aliases(args.aliases)
    for w in warnings:
        print("WARN:", w)
    print(f"Report conditions mapped : {len(cond_map)}")
    print(f"Aliases loaded           : {len(aliases)}  ({args.aliases})")

    # 2. Read manifest, compute new group per pathology.
    with zipfile.ZipFile(args.inp) as zin:
        group_keys = read_group_keys_from_zip(zin)
        manifest = zin.read("manifest.txt").decode("utf-8").replace("\r\n", "\n").replace("\r", "\n")

    if group_keys:
        bad = sorted({g for g in set(cond_map.values()) | set(aliases.values()) if g not in group_keys})
        if bad:
            print(f"WARN: mapping references group keys not in groups.txt: {', '.join(bad)}")

    new_group = {}          # pid -> chosen group key (may equal old, or old if fallback)
    source = {}             # pid -> 'report' | 'fallback-existing' | 'unmapped'
    unmatched_components = collections.Counter()
    old_dist = collections.Counter()
    new_dist = collections.Counter()
    changes = []            # (pid, title, old, new, source)
    n_changed = n_fallback = n_unmapped = 0

    for _, pid, old, title in parse_manifest(manifest):
        old_dist[old or "<none>"] += 1
        g, matched, unmatched = resolve_group(title, cond_map, aliases, priority)
        for u in unmatched:
            unmatched_components[u] += 1
        if g is None:
            # nothing mappable -> keep whatever the record already had
            g = old
            src = "fallback-existing" if old else "unmapped"
            if old:
                n_fallback += 1
            else:
                n_unmapped += 1
        else:
            src = "report"
        new_group[pid] = g
        source[pid] = src
        new_dist[g or "<none>"] += 1
        if (g or None) != (old or None):
            n_changed += 1
            changes.append((pid, title, old or "", g or "", src))

    total = sum(old_dist.values())

    # 3. Report.
    print(f"\nPathologies              : {total}")
    print(f"Regrouped from report    : {sum(1 for s in source.values() if s == 'report')}")
    print(f"Kept existing (no match) : {n_fallback}")
    print(f"Still ungrouped          : {n_unmapped}")
    print(f"Group CHANGED            : {n_changed}")

    def dist_block(title, dist):
        print(f"\n{title}")
        for g, c in dist.most_common():
            print(f"  {g:<14} {c:6d}")
    dist_block("OLD group distribution:", old_dist)
    dist_block("NEW group distribution:", new_dist)

    if unmatched_components:
        print(f"\nUnmatched title components ({len(unmatched_components)} distinct, "
              f"{sum(unmatched_components.values())} occurrences) — add high-frequency ones to "
              f"{os.path.basename(args.aliases)}:")
        for comp, n in unmatched_components.most_common(30):
            print(f"  {n:6d}  {comp!r}")

    report_out = args.report_out or os.path.join(os.path.dirname(os.path.abspath(args.report)),
                                                 "regroup-changes.tsv")
    with open(report_out, "w", encoding="utf-8", newline="\n") as f:
        f.write("id\ttitle\told_group\tnew_group\tsource\n")
        for pid, title, old, new, src in changes:
            f.write(f"{pid}\t{title}\t{old}\t{new}\t{src}\n")
    print(f"\nWrote change report ({len(changes)} rows): {report_out}")

    if not args.out:
        print("\nDry run (no --out): source zip untouched. Re-run with --out to write the regrouped zip.")
        return 0

    # 4. Stream the regrouped zip: rewrite manifest + each changed .dat header, copy the rest.
    print(f"\nWriting regrouped zip -> {args.out}")
    new_manifest_lines = []
    for line in manifest.split("\n"):
        if line.startswith("pathology:") and not line.lstrip().startswith("#"):
            pid = line.split(";", 1)[0].split(":", 1)[1].strip()
            g = new_group.get(pid)
            if g:
                line = set_manifest_group(line, g)
        new_manifest_lines.append(line)
    new_manifest = "\n".join(new_manifest_lines).encode("utf-8")

    written = 0
    with zipfile.ZipFile(args.inp) as zin, \
         zipfile.ZipFile(args.out, "w", zipfile.ZIP_DEFLATED, compresslevel=args.compresslevel) as zout:
        infos = zin.infolist()
        for info in infos:
            name = info.filename
            base = os.path.basename(name)
            if base == "manifest.txt":
                zout.writestr("manifest.txt", new_manifest)
            elif base.endswith(".dat"):
                pid = base[:-4]
                g = new_group.get(pid)
                data = zin.read(name)
                # only decode+rewrite when the group actually changes
                cur_old = None
                # cheap check: does header already carry this group?
                text = data.decode("utf-8")
                if g:
                    head = text.split("\n", 8)
                    has = next((l for l in head if l.startswith("group:")), None)
                    if has != ("group:" + g):
                        text = set_dat_group(text.replace("\r\n", "\n").replace("\r", "\n"), g)
                        data = text.encode("utf-8")
                zout.writestr(base, data)
            else:
                zout.writestr(base, zin.read(name))
            written += 1
            if written % 5000 == 0 or written == len(infos):
                print(f"  {written}/{len(infos)} entries")

    print(f"Done. Wrote {args.out}")
    print("Deployment reminder: to see new groups in an existing install, delete "
          r"%LOCALAPPDATA%\CardioSimulator\pathologies so the app re-extracts.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
